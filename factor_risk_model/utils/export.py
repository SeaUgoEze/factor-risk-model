"""Report export: CSV, Excel (multi-sheet .xlsx), PDF and HTML.

Used by both the interactive app (download buttons) and the CLI (writes to
``outputs/reports/``).  The PDF report is built with matplotlib's
PdfPages - already installed, zero extra dependencies.

Every exporter takes the PipelineResult and a destination and returns
the path it wrote.  All formats share one section order:

    cover / summary -> factor exposures -> portfolio optimization
    -> risk & performance -> stress test -> anomaly detection
"""
from __future__ import annotations

import base64
import html
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages

from factor_risk_model.utils.helpers import fmt_pct

CSV_TABLES = ("exposures", "weights", "risk_summary", "attribution",
              "stress", "anomaly_flags")

# Shared by HTML + PDF: (figure key, section heading, caption, alt text).
# Order is the reading order.
SECTIONS = (
    ("exposures", "1 · Factor exposures",
     "Beta loadings across the universe - the SPY row at the bottom is the "
     "market sanity check (beta ≈ 1.0).",
     "Factor exposures heatmap"),
    ("weights", "2 · Portfolio optimization",
     "Optimal weights under the mandate - target vs achieved exposure.",
     "Optimal portfolio weights"),
    ("cumulative", "3 · Risk & performance - cumulative growth",
     "Growth of one unit: optimal portfolio vs SPY vs equal weight.",
     "Cumulative growth chart"),
    ("drawdowns", "3 · Risk & performance - drawdowns",
     "Peak-to-trough drawdown path: optimal vs SPY vs equal weight.",
     "Drawdown chart"),
    ("attribution", "3 · Risk & performance - factor attribution",
     "Annualized contribution of each factor to portfolio returns.",
     "Factor attribution chart"),
    ("stress", "4 · Stress test",
     "One-month impact of stylized factor shocks, predicted by the factor "
     "model itself.",
     "Stress scenario chart"),
    ("anomaly", "5 · Anomaly detection",
     "Windowed autoencoder reconstruction error on daily portfolio returns; "
     "windows flagged at the 95th-percentile threshold.",
     "Anomaly detection chart"),
)

# Figure key -> table shown beneath it in the same section.
TABLE_UNDER = {
    "exposures": "exposures",
    "weights": "weights",
    "cumulative": "risk_summary",
    "drawdowns": None,
    "attribution": "attribution",
    "stress": "stress",
    "anomaly": "anomaly_flags",
}

TABLE_TITLES = {
    "risk_summary": "Risk summary - optimal vs SPY vs equal weight",
    "exposures": "Factor exposures (betas, 95% CI)",
    "weights": "Optimized portfolio weights",
    "attribution": "Factor attribution (annualized)",
    "stress": "Stress scenarios (one-month factor shocks)",
    "anomaly_flags": "Anomaly detection - flagged windows",
}

# Short labels for the cover's Contents section.  Order matches SECTIONS.
CONTENTS_LABELS = {
    "exposures": "Factor exposures",
    "ci": "Exposures with 95% CI",
    "weights": "Portfolio optimization",
    "cumulative": "Cumulative growth",
    "drawdowns": "Drawdowns",
    "attribution": "Factor attribution",
    "stress": "Stress test",
    "anomaly": "Anomaly detection",
}

# ---- small shared helpers -------------------------------------------------

def _table(result, name):
    """Pull a table out of the result, normalizing dicts to frames."""
    obj = getattr(result, name, None)
    if obj is None:
        return None
    if name == "attribution":
        c = obj["contributions"].copy()
        c["alpha"] = obj["alpha_ann"]
        return c.to_frame("ann_contribution_%").round(4)
    if name == "weights":
        return obj.to_frame("weight")
    if isinstance(obj, dict):
        return pd.DataFrame([obj])
    return obj


def _data_uri(path) -> str:
    """Inline a PNG as a base64 data URI (self-contained HTML report)."""
    b64 = base64.b64encode(Path(path).read_bytes()).decode("ascii")
    return f"data:image/png;base64,{b64}"


def _generated_stamp(generated) -> str:
    """Human 'Report generated …' stamp shared by every report surface."""
    generated = pd.Timestamp(generated)
    return f"Report generated {generated:%Y-%m-%d} at {generated:%H:%M}"


def _headline_metrics(result) -> list[tuple[str, str]]:
    """The four headline numbers (label, value)."""
    m = result.risk_summary.loc["Optimal"]
    return [
        ("Annualized return", fmt_pct(m["ann_return_%"] / 100)),
        ("Annualized vol", fmt_pct(m["ann_vol_%"] / 100)),
        ("Sharpe ratio", f"{m['sharpe']:.2f}"),
        ("Max drawdown", fmt_pct(m["max_drawdown_%"] / 100)),
    ]


def export_csv(result, out_dir: Path) -> list[Path]:
    """One CSV per analysis table."""
    out_dir.mkdir(parents=True, exist_ok=True)
    written = []
    for name in CSV_TABLES:
        frame = _table(result, name)
        if frame is None:
            continue
        path = out_dir / f"{name}.csv"
        frame.to_csv(path)
        written.append(path)
    return written


# ---- Excel ----------------------------------------------------------------

def _summary_rows(result) -> pd.DataFrame:
    """Front sheet: report meta, headline metrics, key findings."""
    rows = [
        ("Report", "Factor-Based Risk & Optimization Model"),
        ("Universe", ", ".join(result.tickers)),
        ("Window", f"{result.start} to {result.end}"),
        ("Factor model", result.factor_model),
        ("Generated", _generated_stamp(pd.Timestamp.now())),
        ("", ""),
        ("HEADLINE METRICS", ""),
        *[(k, v) for k, v in _headline_metrics(result)],
        ("", ""),
        ("KEY FINDINGS", ""),
        *[(f"• {b}", "") for b in result.interpretation()],
    ]
    return pd.DataFrame(rows, columns=["Item", "Detail"])


def export_excel(result, dest) -> object:
    """Single workbook: Summary sheet first, then one sheet per table.

    Every sheet gets a dark header row, frozen top row and auto-sized
    columns.

    ``dest`` is a Path to write, or any writable file-like (e.g. a
    BytesIO for an in-memory download).
    """
    if isinstance(dest, (str, Path)):
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        _summary_rows(result).to_excel(writer, sheet_name="Summary",
                                       index=False)
        for name in CSV_TABLES:
            frame = _table(result, name)
            if frame is not None:
                frame.to_excel(writer, sheet_name=name[:31])
        # ---- style pass over the freshly written workbook ----
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = writer.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="101418",
                                        end_color="101418",
                                        fill_type="solid")
                cell.alignment = Alignment(horizontal="right")
            # the index column holds row labels - keep it left-aligned
            ws["A1"].alignment = Alignment(horizontal="left")
            if ws.title == "Summary":
                ws.column_dimensions["A"].width = 26
                ws.column_dimensions["B"].width = 100
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if cell.value in ("HEADLINE METRICS",
                                          "KEY FINDINGS"):
                            cell.font = Font(bold=True, color="101418")
                            continue
                        if cell.column == 2 or (
                                cell.value and
                                str(cell.value).startswith("•")):
                            cell.alignment = Alignment(wrap_text=True,
                                                       vertical="top")
                continue
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float):
                        cell.number_format = "0.0000"
            for col in ws.columns:
                width = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = \
                    min(max(width + 3, 10), 34)
    return dest


# ---- HTML -----------------------------------------------------------------

_CSS = """
 :root { --ink: #101418; --ink-soft: #3a4148; --accent: #004d00;
         --line: #d7dce2; --head-bg: #101418; --bg: #f7f8fa; }
 * { box-sizing: border-box; }
 body { font-family: Georgia, 'Times New Roman', serif; margin: 0 auto;
        padding: 48px 56px; max-width: 980px; color: var(--ink);
        line-height: 1.55; background: #fff; }
 /* branded cover: 'Factor risk report' in Times New Roman, centered,
    underlined - the border-bottom sits on the shrink-to-fit block so the
    line spans exactly the text width, matching the PDF cover. */
 h1 { margin: 0 auto 12px; color: var(--ink); font-size: 30px;
      font-family: 'Times New Roman', Times, serif; font-weight: 700;
      letter-spacing: 0; text-align: center; width: fit-content;
      padding: 0 2px 7px; border-bottom: 2px solid var(--ink); }
 .meta { color: var(--ink-soft); font-size: 13px; margin: 0 auto 26px;
         text-align: center; border-bottom: 2px solid var(--accent);
         padding-bottom: 14px; }
 .metrics { display: flex; gap: 14px; flex-wrap: wrap; margin: 4px 0 28px; }
 .card { flex: 1 1 170px; border: 1px solid var(--line);
         border-top: 3px solid var(--accent); border-radius: 6px;
         padding: 12px 16px; background: var(--bg); }
 .card .k { font-size: 11px; text-transform: uppercase;
            letter-spacing: .06em; color: var(--ink-soft); }
 .card .v { font-size: 22px; font-weight: 700; margin-top: 4px; }
 h2 { margin: 40px 0 6px; font-size: 19px; color: var(--ink);
      border-left: 4px solid var(--accent); padding-left: 10px; }
 h3 { margin: 26px 0 8px; font-size: 14.5px; color: var(--ink);
      font-weight: 600; letter-spacing: .01em; }
 .lead { margin: 4px 0 18px; font-size: 13.5px; color: var(--ink-soft); }
 figure.chart { margin: 14px 0 22px; }
 figure.chart img { width: 100%; height: auto; border: 1px solid var(--line);
                    border-radius: 6px; }
 figcaption { margin-top: 8px; font-size: 12px; color: var(--ink-soft);
              font-style: italic; }
 table { border-collapse: collapse; width: 100%; font-size: 13px;
         font-family: 'Helvetica Neue', Arial, sans-serif;
         margin-bottom: 8px; }
 th { background: var(--head-bg); color: #fff; font-weight: 600;
      padding: 8px 10px; text-align: right; border: 1px solid var(--head-bg); }
 th:first-child, td:first-child { text-align: left; }
 td { border: 1px solid var(--line); padding: 7px 10px; text-align: right;
      color: var(--ink); }
 tr:nth-child(even) td { background: var(--bg); }
 ul { padding-left: 22px; }
 li { margin: 7px 0; }
 .foot { margin-top: 44px; padding-top: 12px; border-top: 1px solid var(--line);
         font-size: 11px; color: var(--ink-soft); text-align: center; }
 @media print { body { padding: 18px 24px; } h2 { break-after: avoid; }
                table, figure.chart { break-inside: avoid; }
                section { break-before: page; }
                section:first-of-type { break-before: auto; } }
"""


def _round_numeric(frame: pd.DataFrame) -> pd.DataFrame:
    """Round numeric columns only (date/id columns stay untouched)."""
    num = frame.select_dtypes(include="number")
    return frame.assign(**{c: num[c].round(3) for c in num.columns})


def _fig_html(result, key: str, caption: str, alt: str) -> str:
    p = result.figures.get(key)
    if p is None or not Path(p).exists():
        return ""
    return (f'<figure class="chart"><img src="{_data_uri(p)}" '
            f'alt="{html.escape(alt)}">'
            f"<figcaption>{html.escape(caption)}</figcaption></figure>")


def _table_html(result, name: str) -> str:
    frame = _table(result, name)
    if frame is None:
        return ""
    return f"<h3>{TABLE_TITLES[name]}</h3>" \
           f"{_round_numeric(frame).to_html(classes='tbl', border=0)}"


def render_html(result) -> str:
    """The self-contained HTML report as a string (shareable, printable)."""
    cards = "".join(
        f'<div class="card"><div class="k">{html.escape(k)}</div>'
        f'<div class="v">{html.escape(v)}</div></div>'
        for k, v in _headline_metrics(result))

    tail = result.tail
    tail_cards = ""
    if tail:
        tail_items = (("95% VaR (historical)", "var_historical_%"),
                      ("95% VaR (normal)", "var_normal_%"),
                      ("95% CVaR / expected shortfall", "cvar_%"))
        tail_cards = "".join(
            f'<div class="card"><div class="k">{html.escape(k)}</div>'
            f'<div class="v">{html.escape(fmt_pct(tail[v] / 100))}</div></div>'
            for k, v in tail_items if v in tail)

    bullets = "".join(f"<li>{html.escape(b)}</li>"
                      for b in result.interpretation())

    sections = []
    for key, heading, caption, alt in SECTIONS:
        under = TABLE_UNDER.get(key)
        extra = ""
        # the exposures section also carries the 95%-CI chart when present
        if key == "exposures" and "ci" in result.figures:
            extra = _fig_html(result, "ci",
                              "Factor exposures with 95% confidence "
                              "intervals (first ticker).",
                              "Exposure confidence intervals")
        # the optimization section shows the mandate-vs-achieved table
        if key == "weights" and result.optimizer:
            cmp = result.optimizer.get("comparison")
            if cmp is not None:
                extra += (f"<h3>Mandate vs achieved</h3>"
                          f"{_round_numeric(cmp).to_html(classes='tbl', border=0)}")
        tbl = _table_html(result, under) if under else ""
        sections.append(
            f'<section><h2>{html.escape(heading)}</h2>'
            f'<p class="lead">{html.escape(caption)}</p>'
            f'{_fig_html(result, key, caption, alt)}{extra}{tbl}</section>')

    body = "\n".join(sections)
    now = pd.Timestamp.now()
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Factor Risk Model - report</title>
<style>{_CSS}</style></head><body>
<h1>Factor risk report</h1>
<p class="meta">Universe: {html.escape(', '.join(result.tickers))} &middot;
 {html.escape(str(result.start))} to {html.escape(str(result.end))} &middot;
 {html.escape(str(result.factor_model))} &middot;
 {_generated_stamp(now)}</p>
<section class="summary">
<h2>Executive summary</h2>
<p class="lead">Headline risk-adjusted metrics for the optimized portfolio.</p>
{cards}
<h2>Key findings</h2><ul>{bullets}</ul>
</section>
{body}
<section class="tailrisk">
<h2>Tail risk</h2>
<p class="lead">95th-percentile downside measures on daily returns.</p>
<div class="metrics">{tail_cards}</div>
</section>
<p class="foot">© {pd.Timestamp.now():%Y} Factor Risk Model · runs fully locally ·
 data: Yahoo Finance &amp; Kenneth French Data Library</p>
</body></html>"""


def export_html(result, dest) -> object:
    """Write the HTML report to a Path or a writable file-like."""
    html_text = render_html(result)
    if isinstance(dest, (str, Path)):
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_text(html_text, encoding="utf-8")
    else:
        dest.write(html_text.encode("utf-8"))
    return dest


# ---- PDF ------------------------------------------------------------------

INK = "#101418"        # near-black body text
INK_SOFT = "#3a4148"    # muted secondary text
NAVY = "#101418"        # header fill - white bold text on top
ACCENT = "#004d00"


def _style_table(tbl, n_rows: int, n_cols: int) -> None:
    """matplotlib table -> dark header row, white bold header text,
    zebra body rows, near-black body text (row labels live in col -1)."""
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8.5)
    tbl.scale(1, 1.2)
    for j in range(n_cols):
        cell = tbl[(0, j)]
        cell.set_facecolor(NAVY)
        cell.set_text_props(color="white", fontweight="bold")
    for i in range(n_rows):
        for j in range(n_cols):
            cell = tbl[(i + 1, j)]
            cell.set_facecolor("#f5f7fa" if i % 2 == 0 else "white")
            cell.set_text_props(color=INK)
        label = tbl[(i + 1, -1)]
        label.set_facecolor("#e8edf5")
        label.set_text_props(color=INK, fontweight="bold")


def _chart_page(pdf, img, title: str, caption: str, page_no: int,
                generated) -> None:
    """One chart page: title band + caption + aspect-fitted image + footer."""
    h, w = img.shape[:2]
    aspect = w / h

    fig = plt.figure(figsize=(8.5, 11))
    fig.patch.set_facecolor("white")

    # title band
    fig.text(0.5, 0.965, title, ha="center", va="center", fontsize=16,
             fontweight="bold", color=INK)
    fig.text(0.5, 0.925, caption, ha="center", va="center", fontsize=9.5,
             color=INK_SOFT)
    fig.add_artist(plt.Line2D([0.14, 0.86], [0.90, 0.90], color=ACCENT,
                              lw=1.4, transform=fig.transFigure))

    # fit the image inside the area below the band, keeping proportions
    area_w, area_h, y0 = 0.92, 0.83, 0.035
    if aspect >= area_w / area_h:          # width-limited
        iw, ih = area_w, area_w / aspect
    else:                                   # height-limited
        iw, ih = area_h * aspect, area_h
    ix, iy = (1 - iw) / 2, y0 + (area_h - ih) / 2
    ax = fig.add_axes([ix, iy, iw, ih])
    ax.imshow(img)
    ax.axis("off")

    fig.text(0.5, 0.012,
             f"Factor Risk Model · page {page_no} · "
             f"{_generated_stamp(generated)} · runs fully locally",
             ha="center", va="bottom", fontsize=8, color=INK_SOFT)
    pdf.savefig(fig)
    plt.close(fig)


def export_pdf(result, figures: dict[str, Path], dest) -> object:
    """Matplotlib PdfPages: cover with Contents + one titled page per chart.

    The cover doubles as a table of contents, listing every chart page
    with its true page number.  ``dest`` is a Path to write, or a writable
    file-like (BytesIO).
    """
    if isinstance(dest, (str, Path)):
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
    summary = result.risk_summary.round(2)
    generated = pd.Timestamp.now()

    # Resolve the ordered chart-page list FIRST, so the cover's Contents
    # section can print the true page number of every chart page.
    order = [key for key, *_ in SECTIONS]
    if "ci" in figures:
        order.insert(order.index("exposures") + 1, "ci")
    chart_pages = []   # (figure key, page title, caption, image array)
    for key in order:
        p = figures.get(key)
        if p is None or not Path(p).exists():
            continue
        try:
            img = plt.imread(p)
        except Exception:
            continue   # skip a corrupt/unreadable chart, don't abort
        if key == "ci":
            title = "Factor exposures with 95% confidence intervals"
            caption = "Point estimates with 95% CI for the first ticker."
        else:
            heading = next(h for k, h, *_ in SECTIONS if k == key)
            title = heading.split("·", 1)[-1].strip()
            caption = next(c for k, _, c, _ in SECTIONS if k == key)
        chart_pages.append((key, title, caption, img))

    page_no = 1

    with PdfPages(dest) as pdf:
        # ---- cover / metrics page ----
        fig = plt.figure(figsize=(8.5, 11))
        fig.patch.set_facecolor("white")
        ax = fig.add_axes([0.08, 0.05, 0.84, 0.90])
        ax.axis("off")
        # Pin the data limits: drawing transAxes rules/leaders below would
        # otherwise autoscale them to a tiny range and push every data-
        # coordinate text (title, meta, contents, key findings) off-page.
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.autoscale(False)
        # Title: 'Factor risk report' in Times New Roman, centered and
        # underlined (the rule is drawn from the measured text width).
        ttl = ax.text(0.5, 0.97, "Factor risk report", ha="center",
                      va="top", fontsize=22, fontweight="bold",
                      fontfamily=["Times New Roman", "serif"], color=INK)
        ax.text(0.5, 0.92,
                f"{result.start} to {result.end}  ·  {result.factor_model}  "
                f"·  {len(result.tickers)} names",
                ha="center", va="top", fontsize=11, color=INK_SOFT)
        ax.plot([0.0, 1.0], [0.89, 0.89], color=ACCENT, lw=1.2,
                transform=ax.transAxes)
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        te = ttl.get_window_extent(renderer)
        ae = ax.get_window_extent(renderer)
        ux0 = (te.x0 - ae.x0) / ae.width
        ux1 = (te.x1 - ae.x0) / ae.width
        uy = (te.y0 - ae.y0) / ae.height - 5 / ae.height
        ax.plot([ux0, ux1], [uy, uy], color=INK, lw=1.3,
                transform=ax.transAxes)
        # ---- Contents: every chart page with its page number ----
        # Two columns of entries with dotted leaders; the leaders are drawn
        # from the measured end of each label to the measured start of its
        # page number, so no line ever crosses the text.
        if chart_pages:
            n_rows = (len(chart_pages) + 1) // 2
            row_h, col_w, col_gap = 0.034, 0.44, 0.10
            ax.text(0.02, 0.868, "Contents", va="top", ha="left",
                    fontsize=13, fontweight="bold", color=INK)
            first_y = 0.845
            entries = []
            for i, (key, _title, _cap, _img) in enumerate(chart_pages):
                col, row = divmod(i, n_rows)
                x0 = 0.02 + col * (col_w + col_gap)
                y = first_y - row * row_h
                label = CONTENTS_LABELS.get(key, _title)
                t = ax.text(x0, y, label, va="center", ha="left",
                            fontsize=9.5, color=INK)
                num = ax.text(x0 + col_w, y, str(2 + i), va="center",
                              ha="right", fontsize=9.5, color=INK_SOFT,
                              fontweight="bold")
                entries.append((t, num, y))
            fig.canvas.draw()
            renderer = fig.canvas.get_renderer()
            inv = ax.transAxes.inverted()
            for t, num, y in entries:
                te = t.get_window_extent(renderer)
                ne = num.get_window_extent(renderer)
                ym = (te.y0 + te.y1) / 2
                xs = inv.transform((te.x1 + 3, ym))[0]
                xe = inv.transform((ne.x0 - 3, ym))[0]
                if xe > xs:
                    ax.plot([xs, xe], [y, y], color="#aab2bc", lw=0.7,
                            ls=(0, (1, 2)), transform=ax.transAxes)
            last_y = first_y - (n_rows - 1) * row_h
            table_top = last_y - 0.05
        else:
            table_top = 0.85
        # Risk summary table - explicit proportional column widths so every
        # column header and value is fully visible on the page.
        cols = list(summary.columns)
        lens = [max(len(str(c)), 3) for c in cols]
        col_widths = [0.14] + [0.86 * (l / sum(lens)) for l in lens]
        tbl = ax.table(cellText=summary.round(3).values,
                       colLabels=cols, rowLabels=summary.index,
                       cellLoc="right", loc="upper center",
                       bbox=[0.02, table_top - 0.33, 0.96, 0.33],
                       colWidths=col_widths)
        _style_table(tbl, len(summary.index), len(cols))
        bullets = "\n".join(f"  • {b}" for b in result.interpretation())
        # Measure the table's REAL extent: matplotlib tables can grow past
        # their hint bbox (previously the table ran into the footer text).
        # Everything below the table is placed from its measured bottom
        # edge, so no overlap is possible.
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        try:
            te = tbl.get_window_extent(renderer)
            ae = ax.get_window_extent(renderer)
            table_bottom = (te.y0 - ae.y0) / (ae.y1 - ae.y0)
        except Exception:
            table_bottom = 0.42
        table_bottom = max(table_bottom, 0.30)
        kf_y = table_bottom - 0.045
        ax.text(0.02, kf_y, "Key findings", va="top", ha="left",
                fontsize=12, fontweight="bold", color=INK)
        ax.text(0.02, kf_y - 0.035, bullets, va="top", ha="left",
                fontsize=9.5, family="monospace", color=INK)
        # The 'Report generated ...' stamp sits at the bottom of every page
        # (cover and chart pages) - far below the table's measured bottom
        # edge, so overlap is impossible.
        fig.text(0.5, 0.012,
                 f"Factor Risk Model · page {page_no} · "
                 f"{_generated_stamp(generated)} · runs fully locally",
                 ha="center", va="bottom", fontsize=8, color=INK_SOFT)
        pdf.savefig(fig)
        plt.close(fig)

        # ---- one page per chart, in reading order (matching the
        #      Contents page numbers) ----
        for _key, title, caption, img in chart_pages:
            page_no += 1
            _chart_page(pdf, img, title, caption, page_no, generated)
    return dest
